from openpyxl import load_workbook


async def read_excel(file_path: str):
    workbook = load_workbook(file_path, read_only=True)
    sheet = workbook.active
    return sheet


async def parse_excel_rows(sheet):
    cases = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        competency = str(row[0])
        rol = str(row[1])
        year = str(row[2])
        court = None
        book = None
        if competency == "Corte Apelaciones":
            court = str(row[3])
            book = str(row[4])
        validate_row_data(competency, rol, year, court, book)
        cases.append(
            {
                "competency": competency,
                "rol": rol,
                "year": year,
                "court": court,
                "book": book,
            }
        )
    return cases


def validate_row_data(
    competency: str, rol: str, year: str, court: str | None, book: str | None
):

    COMPETENCY_OPTIONS = [
        "Corte Suprema",
        "Corte Apelaciones",
        "Civil",
        "Laboral",
        "Penal",
        "Cobranza",
        "Familia",
    ]
    COURTS_OPTIONS = [
        "Todos",
        "C.A. de Arica",
        "C.A. de Iquique",
        "C.A. de Antofagasta",
        "C.A. de Copiapó",
        "C.A. de La Serena",
        "C.A. de Valparaíso",
        "C.A. de Rancagua",
        "C.A. de Talca",
        "C.A. de Chillan",
        "C.A. de Concepción",
        "C.A. de Temuco",
        "C.A. de Valdivia",
        "C.A. de Puerto Montt",
        "C.A. de Coyhaique",
        "C.A. de Punta Arenas",
        "C.A. de Santiago",
        "C.A. de San Miguel",
    ]
    BOOKS_OPTIONS = [
        "Todos",
        "Civil",
        "Familia",
        "Laboral - Cobranza",
        "Penal",
        "Contencioso Administrativo",
        "Tributario Y Aduanero",
        "Protección",
        "Amparo",
        "Policia Local",
        "Exhorto",
        "Ley De Navegación",
        "Ambiental",
        "Traspaso Corte Marcial",
        "Ministro Primera Instancia Y Fuero",
        "Com. Lib. Cond.",
    ]

    if not competency or competency not in COMPETENCY_OPTIONS:
        raise ValueError(f"Competency {competency} is not valid")
    if not rol:
        raise ValueError(f"Rol {rol} is not valid")
    if not year:
        raise ValueError(f"Year {year} is not valid")
    if competency == "Corte Apelaciones" and (not court or court not in COURTS_OPTIONS):
        raise ValueError(f"Court {court} is not valid")
    if competency == "Corte Apelaciones" and (not book or book not in BOOKS_OPTIONS):
        raise ValueError(f"Book {book} is not valid")
    return True
